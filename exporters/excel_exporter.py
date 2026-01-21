"""
엑셀 파일 생성 모듈 (v4 - 완전 한글화, 대폭 확장)
- 기업명 포함
- 계정과목 설명 시트 (50개+)
- 전문 활용 가이드 시트
- 전체 컬럼 한글화
- 제작자/일시 표시
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Dict, Optional
from datetime import datetime
import logging
import os

# 계정 설명 가져오기
try:
    from config.account_explanations import ACCOUNT_EXPLANATIONS
except ImportError:
    ACCOUNT_EXPLANATIONS = {}

logger = logging.getLogger("kr_stock_collector.exporter")


class ExcelExporter:
    """엑셀 파일 생성 클래스 (완전 한글화)"""
    
    # 스타일
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
    HEADER_FILL = PatternFill('solid', fgColor='4472C4')
    ALT_FILL = PatternFill('solid', fgColor='F2F2F2')
    BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 전체 컬럼 한글화 매핑
    COLUMN_KOREAN = {
        # 기본 정보
        'stock_code': '종목코드', 'Code': '종목코드', 'Name': '기업명',
        'Market': '시장', 'Sector': '업종', 'Industry': '산업',
        'market_cap': '시가총액', 'shares': '상장주식수', 'date': '기준일',
        
        # 주가
        'open': '시가', 'high': '고가', 'low': '저가', 'close': '종가',
        'volume': '거래량', 'value': '거래대금', 'change': '등락률',
        
        # 투자지표
        'bps': 'BPS', 'per': 'PER', 'pbr': 'PBR', 'eps': 'EPS',
        'div_yield': '배당수익률', 'dps': 'DPS',
        
        # 재무제표 (OpenDART)
        'corp_code': '기업코드', 'corp_name': '기업명', 'bsns_year': '사업연도',
        'reprt_code': '보고서', 'account_nm': '계정과목',
        'thstrm_amount': '당기금액', 'frmtrm_amount': '전기금액',
        'bfefrmtrm_amount': '전전기금액', 'fs_div': '재무제표구분',
        'fs_nm': '재무제표명', 'sj_div': '계정구분', 'sj_nm': '계정분류',
        'thstrm_nm': '당기명', 'thstrm_dt': '당기일자',
        'frmtrm_nm': '전기명', 'frmtrm_dt': '전기일자',
        'bfefrmtrm_nm': '전전기명', 'bfefrmtrm_dt': '전전기일자',
        'ord': '순서', 'currency': '통화',
        
        # 거시경제
        'TIME': '날짜', 'DATA_VALUE': '값', 'STAT_NAME': '통계명',
        'indicator': '지표명', 'category': '카테고리', 'source': '출처',
        'series_id': '시리즈ID',
    }
    
    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.stock_names = {}
        self.created_time = datetime.now()
    
    def set_stock_names(self, stock_list: pd.DataFrame) -> None:
        """종목명 매핑 설정"""
        if stock_list is not None and 'Code' in stock_list.columns and 'Name' in stock_list.columns:
            self.stock_names = dict(zip(stock_list['Code'], stock_list['Name']))
    
    def _add_company_name(self, df: pd.DataFrame, code_col: str = 'stock_code') -> pd.DataFrame:
        """기업명 컬럼 추가"""
        if code_col in df.columns and self.stock_names:
            name_col = df[code_col].map(self.stock_names)
            idx = df.columns.get_loc(code_col) + 1
            df.insert(idx, '기업명', name_col)
        return df
    
    def _korean_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """모든 컬럼을 한글로 변환"""
        rename_map = {k: v for k, v in self.COLUMN_KOREAN.items() if k in df.columns}
        return df.rename(columns=rename_map)
    
    def _auto_width(self, ws, min_w: int = 8, max_w: int = 35) -> None:
        """컬럼 너비 자동"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        val = str(cell.value)
                        length = sum(1.5 if '\uac00' <= c <= '\ud7a3' else 1 for c in val)
                        max_len = max(max_len, length)
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)
    
    def _apply_table_style(self, ws, header_row: int = 1) -> None:
        """테이블 스타일 적용"""
        for cell in ws[header_row]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=1):
            for cell in row:
                cell.border = self.BORDER
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_FILL
                if isinstance(cell.value, (int, float)):
                    if abs(cell.value) >= 1000:
                        cell.number_format = '#,##0'
                    elif cell.value != 0 and abs(cell.value) < 100 and cell.value != int(cell.value):
                        cell.number_format = '0.00'
    
    def add_usage_guide_sheet(self) -> None:
        """📚 전문 활용 가이드 시트"""
        ws = self.wb.create_sheet("📚 활용가이드", 0)
        
        content = [
            ("══════════════════════════════════════════════════════════════════", "", ""),
            ("📊 충북대학교 가치투자학회 종목 스크리닝 시스템", "", ""),
            (f"   제작자: 이찬희  |  생성일시: {self.created_time.strftime('%Y-%m-%d %H:%M')}", "", ""),
            ("══════════════════════════════════════════════════════════════════", "", ""),
            ("", "", ""),
            
            # 1. 스크리닝 전략
            ("━━━ 📈 1. 투자 스크리닝 전략 ━━━", "", ""),
            ("", "", ""),
            ("【 벤저민 그레이엄 스타일 (안전마진 투자) 】", "", ""),
            ("", "투자지표 시트 → 필터 적용:", ""),
            ("", "  • PER < 10 (저평가)", ""),
            ("", "  • PBR < 1 (청산가치 이하)", ""),
            ("", "  • 배당수익률 > 3%", ""),
            ("", "  → 결과: 안전마진이 큰 저평가 우량주", ""),
            ("", "", ""),
            ("【 워렌 버핏 스타일 (경쟁우위 투자) 】", "", ""),
            ("", "재무제표 시트 → 필터 적용:", ""),
            ("", "  • ROE > 15% (자기자본이익률)", ""),
            ("", "  • 영업이익률 > 10%", ""),
            ("", "  • 부채비율 < 50%", ""),
            ("", "  → 결과: 지속적 경쟁우위 보유 우량 기업", ""),
            ("", "", ""),
            ("【 피터 린치 스타일 (성장주 발굴) 】", "", ""),
            ("", "재무제표 시트에서:", ""),
            ("", "  • 매출성장률 > 20%", ""),
            ("", "  • 영업이익성장률 > 20%", ""),
            ("", "  • PEG < 1 (PER / 성장률)", ""),
            ("", "  → 결과: 저평가된 고성장 중소형주", ""),
            ("", "", ""),
            ("【 배당 투자 전략 】", "", ""),
            ("", "투자지표 시트에서:", ""),
            ("", "  • 배당수익률 > 4%", ""),
            ("", "  • 배당성향 20~60% (적정 범위)", ""),
            ("", "  • 5년 연속 배당 지급", ""),
            ("", "  → 결과: 안정적 현금흐름 배당주", ""),
            ("", "", ""),
            
            # 2. 재무분석 활용법
            ("━━━ 💰 2. 재무제표 분석 가이드 ━━━", "", ""),
            ("", "", ""),
            ("【 수익성 분석 (얼마나 잘 버는가?) 】", "", ""),
            ("지표", "산식", "기준"),
            ("매출총이익률", "= 매출총이익 / 매출액", "30%+ 양호"),
            ("영업이익률", "= 영업이익 / 매출액", "10%+ 우량"),
            ("순이익률", "= 당기순이익 / 매출액", "5%+ 양호"),
            ("ROE", "= 당기순이익 / 자본총계", "15%+ 우수"),
            ("ROA", "= 당기순이익 / 자산총계", "5%+ 양호"),
            ("", "", ""),
            ("【 안정성 분석 (얼마나 안전한가?) 】", "", ""),
            ("부채비율", "= 부채총계 / 자본총계", "100% 이하 안정"),
            ("유동비율", "= 유동자산 / 유동부채", "100%+ 양호"),
            ("당좌비율", "= (유동자산-재고) / 유동부채", "100%+ 우수"),
            ("이자보상배율", "= 영업이익 / 이자비용", "3배+ 안전"),
            ("", "", ""),
            ("【 성장성 분석 (얼마나 성장하는가?) 】", "", ""),
            ("매출성장률", "= (당기매출-전기매출) / 전기매출", "10%+ 성장"),
            ("영업이익성장률", "= (당기영업이익-전기) / 전기", "10%+ 성장"),
            ("자산성장률", "= (당기자산-전기자산) / 전기자산", "양호"),
            ("", "", ""),
            
            # 3. 거시경제 활용
            ("━━━ 🌍 3. 거시경제 지표 활용법 ━━━", "", ""),
            ("", "", ""),
            ("【 금리와 주식시장 】", "", ""),
            ("금리 인상기", "→ 가치주, 금융주, 현금보유 기업 유리", ""),
            ("금리 인하기", "→ 성장주, 기술주, 부채 많은 기업 유리", ""),
            ("", "", ""),
            ("【 주요 지표 해석 】", "", ""),
            ("VIX > 30", "→ 시장 공포 극대화, 매수 기회 검토", ""),
            ("10Y-2Y 스프레드 역전", "→ 경기침체 신호, 방어주 비중 확대", ""),
            ("달러인덱스 상승", "→ 신흥국 주식 부담, 수출주 주의", ""),
            ("유가 급등", "→ 인플레이션 우려, 에너지주 관심", ""),
            ("비트코인 급등", "→ 위험자산 선호, 성장주 동반 상승 가능", ""),
            ("", "", ""),
            
            # 4. 취업 활용
            ("━━━ 💼 4. 취업/커리어 활용법 ━━━", "", ""),
            ("", "", ""),
            ("【 금융권 면접 활용 】", "", ""),
            ("", "• '이 시스템으로 2,500개 기업 재무데이터 분석 경험'", ""),
            ("", "• 'OpenDART, FRED API를 활용한 데이터 수집 자동화'", ""),
            ("", "• 'Python으로 스크리닝 시스템 개발 및 운영'", ""),
            ("", "", ""),
            ("【 리서치 역량 어필 】", "", ""),
            ("", "• '데이터 기반 투자 논리 전개 능력'", ""),
            ("", "• '재무제표 분석을 통한 기업가치 평가'", ""),
            ("", "• '거시경제 지표와 주식시장 상관관계 분석'", ""),
            ("", "", ""),
            ("【 포트폴리오 레포트 작성 예시 】", "", ""),
            ("", "1. 투자 아이디어 발굴 (스크리닝)", ""),
            ("", "2. 재무분석 (수익성/안정성/성장성)", ""),
            ("", "3. Valuation (PER/PBR/DCF)", ""),
            ("", "4. 리스크 분석", ""),
            ("", "5. 투자 결론 및 목표가", ""),
            ("", "", ""),
            
            # 5. 시트별 가이드
            ("━━━ 📑 5. 각 시트 사용법 ━━━", "", ""),
            ("", "", ""),
            ("시트명", "내용", "필터 팁"),
            ("📋 종목리스트", "전체 종목 코드/기업명/시장", "시장=KOSPI로 필터"),
            ("📑 재무제표", "3년치 재무제표 데이터", "계정과목=매출액으로 필터"),
            ("📈 투자지표", "PER/PBR/배당률 등", "PER<10 AND PBR<1"),
            ("💹 주가", "시가/고가/저가/종가", "특정 종목 필터"),
            ("🌍 거시경제", "40개+ 글로벌 지표", "카테고리로 필터"),
            ("📖 계정설명", "계정과목 한글 설명", "모르는 계정 검색"),
            ("", "", ""),
            
            # 6. 주의사항
            ("━━━ ⚠️ 6. 투자 주의사항 ━━━", "", ""),
            ("", "", ""),
            ("", "⚠️ 과거 실적이 미래 성과를 보장하지 않습니다", ""),
            ("", "⚠️ 업종별 적정 수치가 다릅니다 (비교 필수)", ""),
            ("", "⚠️ 일회성 손익(자산매각 등)을 반드시 확인하세요", ""),
            ("", "⚠️ 투자 결정 전 애널리스트 리포트도 참고하세요", ""),
            ("", "⚠️ 본 자료는 투자 권유가 아닙니다", ""),
            ("", "", ""),
            ("══════════════════════════════════════════════════════════════════", "", ""),
        ]
        
        for idx, (col1, col2, col3) in enumerate(content, 1):
            ws.cell(row=idx, column=1, value=col1)
            ws.cell(row=idx, column=2, value=col2)
            ws.cell(row=idx, column=3, value=col3)
            
            if col1.startswith("📊"):
                ws.cell(row=idx, column=1).font = Font(bold=True, size=14, color='1F4E79')
            elif col1.startswith(("━━━", "══")):
                ws.cell(row=idx, column=1).font = Font(bold=True, size=11, color='4472C4')
            elif col1.startswith("【"):
                ws.cell(row=idx, column=1).font = Font(bold=True, size=10)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
    
    def add_account_explanation_sheet(self) -> None:
        """📖 계정과목 설명 시트"""
        ws = self.wb.create_sheet("📖 계정설명")
        
        headers = ['계정명', '영문명', '분류', '설명', '활용법', '주의사항']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        
        row = 2
        for name, info in ACCOUNT_EXPLANATIONS.items():
            ws.cell(row=row, column=1, value=info.get('한글명', name))
            ws.cell(row=row, column=2, value=info.get('영문명', ''))
            ws.cell(row=row, column=3, value=info.get('분류', ''))
            ws.cell(row=row, column=4, value=info.get('설명', ''))
            ws.cell(row=row, column=5, value=info.get('활용', ''))
            ws.cell(row=row, column=6, value=info.get('주의사항', ''))
            
            if row % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = self.ALT_FILL
            row += 1
        
        ws.auto_filter.ref = f"A1:F{row-1}"
        self._auto_width(ws)
        ws.freeze_panes = 'B2'
    
    def add_summary_sheet(self, summary: Dict) -> None:
        """요약 시트"""
        ws = self.wb.create_sheet("📊 요약", 1)
        
        ws['A1'] = "📊 수집 결과 요약"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        data = [
            ('생성일시', self.created_time.strftime('%Y-%m-%d %H:%M:%S')),
            ('제작자', '이찬희 (충북대학교 가치투자학회)'),
            ('', ''),
            ('총 종목 수', f"{summary.get('total_stocks', 0):,}개"),
            ('재무제표', f"{summary.get('financial_count', 0):,}건"),
            ('투자지표', f"{summary.get('indicator_count', 0):,}건"),
            ('주가 데이터', f"{summary.get('price_count', 0):,}건"),
            ('거시경제', f"{summary.get('macro_count', 0):,}건"),
        ]
        
        for idx, (label, value) in enumerate(data, start=3):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
    
    def add_stock_list_sheet(self, df: pd.DataFrame, cap_df: pd.DataFrame = None) -> None:
        """종목리스트 시트"""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("📋 종목리스트")
        
        if cap_df is not None and not cap_df.empty:
            if 'stock_code' in cap_df.columns:
                cap_df = cap_df.rename(columns={'stock_code': 'Code'})
            df = df.merge(cap_df, on='Code', how='left')
        
        df = self._korean_columns(df)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_table_style(ws)
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)
        ws.freeze_panes = 'C2'
        
        logger.info(f"종목리스트 시트: {len(df)}건")
    
    def add_financial_sheet(self, df: pd.DataFrame) -> None:
        """재무제표 시트"""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("📑 재무제표")
        
        # 기업명 추가
        if 'corp_name' not in df.columns and 'stock_code' in df.columns:
            df = self._add_company_name(df.copy(), 'stock_code')
        
        # 중요 컬럼 순서 정리
        priority = ['stock_code', '기업명', 'corp_name', 'bsns_year', 'account_nm', 
                    'thstrm_amount', 'frmtrm_amount', 'bfefrmtrm_amount']
        ordered = [c for c in priority if c in df.columns]
        others = [c for c in df.columns if c not in priority]
        df = df[ordered + others]
        
        # 컬럼 한글화
        df = self._korean_columns(df)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_table_style(ws)
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)
        ws.freeze_panes = 'D2'
        
        logger.info(f"재무제표 시트: {len(df)}건")
    
    def add_indicator_sheet(self, df: pd.DataFrame) -> None:
        """투자지표 시트"""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("📈 투자지표")
        
        df = self._add_company_name(df.copy(), 'stock_code')
        df = self._korean_columns(df)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_table_style(ws)
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)
        ws.freeze_panes = 'C2'
        
        logger.info(f"투자지표 시트: {len(df)}건")
    
    def add_price_sheet(self, df: pd.DataFrame) -> None:
        """주가 시트"""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("💹 주가")
        
        df = self._add_company_name(df.copy(), 'stock_code')
        df = self._korean_columns(df)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_table_style(ws)
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)
        ws.freeze_panes = 'C2'
        
        logger.info(f"주가 시트: {len(df)}건")
    
    def add_macro_sheet(self, df: pd.DataFrame) -> None:
        """거시경제 시트"""
        if df.empty:
            return
        
        ws = self.wb.create_sheet("🌍 거시경제")
        
        df = self._korean_columns(df)
        
        keep_cols = ['날짜', '카테고리', '지표명', '값', '출처', '시리즈ID']
        available = [c for c in keep_cols if c in df.columns]
        if available:
            df = df[available]
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        self._apply_table_style(ws)
        ws.auto_filter.ref = ws.dimensions
        self._auto_width(ws)
        ws.freeze_panes = 'B2'
        
        logger.info(f"거시경제 시트: {len(df)}건")
    
    def save(self, filename: str = None) -> str:
        """저장"""
        if filename is None:
            timestamp = self.created_time.strftime('%Y%m%d_%H%M%S')
            filename = f"종목스크리너_{timestamp}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        self.wb.save(filepath)
        logger.info(f"엑셀 파일 저장: {filepath}")
        return filepath
    
    def export_all(
        self,
        financial_data: pd.DataFrame = None,
        price_data: pd.DataFrame = None,
        indicator_data: pd.DataFrame = None,
        macro_data: pd.DataFrame = None,
        stock_list: pd.DataFrame = None,
        market_cap_df: pd.DataFrame = None,
        filename: str = None
    ) -> str:
        """전체 내보내기"""
        
        self.set_stock_names(stock_list)
        
        summary = {
            'timestamp': self.created_time.strftime('%Y-%m-%d %H:%M:%S'),
            'total_stocks': len(stock_list) if stock_list is not None else 0,
            'financial_count': len(financial_data) if financial_data is not None else 0,
            'price_count': len(price_data) if price_data is not None else 0,
            'indicator_count': len(indicator_data) if indicator_data is not None else 0,
            'macro_count': len(macro_data) if macro_data is not None else 0,
        }
        
        # 시트 추가 (순서대로)
        self.add_usage_guide_sheet()
        self.add_summary_sheet(summary)
        
        if stock_list is not None and not stock_list.empty:
            self.add_stock_list_sheet(stock_list, market_cap_df)
        
        if financial_data is not None and not financial_data.empty:
            self.add_financial_sheet(financial_data)
        
        if indicator_data is not None and not indicator_data.empty:
            self.add_indicator_sheet(indicator_data)
        
        if price_data is not None and not price_data.empty:
            self.add_price_sheet(price_data)
        
        if macro_data is not None and not macro_data.empty:
            self.add_macro_sheet(macro_data)
        
        # 계정 설명 시트 (마지막)
        self.add_account_explanation_sheet()
        
        return self.save(filename)
